"""Admin — research study management routes."""
import csv
import io
import json
import zipfile
from datetime import datetime, timezone

from flask import render_template, redirect, url_for, flash, request, current_app
from flask import Response
from flask_login import current_user

from app.blueprints.admin import admin_bp
from app.extensions import db
from app.models.user import User
from app.models.task import Task, TaskSubmission
from app.models.settings import Notification
from app.utils.decorators import admin_required
from app.utils.audit import log_action


# ── Private helpers ────────────────────────────────────────────────────────────

def _set_study_dates(study) -> None:
    """Parse date/time fields from form and set on study object.

    The browser submits datetime-local values in the user's local time.
    A hidden field 'utc_offset_minutes' carries getTimezoneOffset() (positive
    means behind UTC, e.g. UTC+2 → -120).  We subtract that offset to get UTC.
    """
    from datetime import timedelta
    try:
        offset_min = int(request.form.get('utc_offset_minutes', 0) or 0)
    except (ValueError, TypeError):
        offset_min = 0

    def _parse_dt(key):
        val = request.form.get(key, '').strip()
        if not val:
            return None
        try:
            # Parse as naive local time, shift to UTC
            local_dt = datetime.fromisoformat(val)
            utc_dt = local_dt + timedelta(minutes=offset_min)
            return utc_dt.replace(tzinfo=timezone.utc)
        except Exception:
            return None

    study.enrollment_start = _parse_dt('enrollment_start')
    study.enrollment_end = _parse_dt('enrollment_end')
    study.task_start = _parse_dt('task_start')
    study.task_end = _parse_dt('task_end')
    try:
        study.max_participants = int(request.form.get('max_participants', '') or 0) or None
    except (ValueError, TypeError):
        study.max_participants = None


def _save_study_steps(study) -> None:
    """Rebuild study steps from JSON posted as 'steps_json'."""
    from app.models.study import StudyStep
    raw = request.form.get('steps_json', '[]')
    try:
        steps_data = json.loads(raw)
    except Exception:
        steps_data = []

    StudyStep.query.filter_by(study_id=study.id).delete()
    db.session.flush()

    def _parse_step_dt(val):
        if not val:
            return None
        try:
            return datetime.fromisoformat(val).replace(tzinfo=timezone.utc)
        except Exception:
            return None

    for idx, s in enumerate(steps_data):
        step = StudyStep(
            study_id=study.id,
            step_order=idx,
            step_type=s.get('type', 'task'),
            survey_id=int(s['survey_id']) if s.get('survey_id') else None,
            task_id=s.get('task_id') or None,
            label=s.get('label', '').strip() or None,
            wave_number=int(s.get('wave_number') or 0),
            available_from=_parse_step_dt(s.get('available_from')),
            available_until=_parse_step_dt(s.get('available_until')),
            allow_late_submission=bool(s.get('allow_late_submission')),
            late_penalty_note=s.get('late_penalty_note', '').strip() or None,
            condition_id=int(s['condition_id']) if s.get('condition_id') else None,
            condition_ids=json.dumps([int(x) for x in s['condition_ids']]) if s.get('condition_ids') else None,
            available_agents=json.dumps(s['available_agents']) if isinstance(s.get('available_agents'), list) else (s.get('available_agents') or None),
            agent_choice_intro=s.get('agent_choice_intro', '').strip() or None,
        )
        db.session.add(step)


def _build_tracking_config() -> str:
    """Build JSON tracking_config from the current POST request form data.

    Three UI groups map to underlying event types:
      tracking_cursor -> mousemove, click, chat_focus
      tracking_bpmn   -> bpmn_add, bpmn_remove, bpmn_move, bpmn_connect, bpmn_disconnect, bpmn_rename, ai_action
      tracking_chat   -> chat_message, ai_action, llm_prompt
    """
    enabled = bool(request.form.get('tracking_enabled'))
    events = []
    if request.form.get('tracking_cursor'):
        events += ['mousemove', 'click', 'chat_focus']
    if request.form.get('tracking_bpmn'):
        events += ['bpmn_add', 'bpmn_remove', 'bpmn_move', 'bpmn_connect', 'bpmn_disconnect', 'bpmn_rename', 'ai_action']
    if request.form.get('tracking_chat'):
        events += ['chat_message', 'ai_action', 'llm_prompt']
    # Deduplicate while preserving order (ai_action may appear in both bpmn+chat groups)
    seen = set()
    unique_events = []
    for e in events:
        if e not in seen:
            seen.add(e)
            unique_events.append(e)
    return json.dumps({'enabled': enabled, 'events': unique_events})


# ── Studies CRUD ──────────────────────────────────────────────────────────────

@admin_bp.route('/studies')
@admin_required
def studies_list():
    from app.models.study import Study
    show_archived = request.args.get('show_archived', '0') == '1'
    q = Study.query
    if not show_archived:
        q = q.filter_by(is_archived=False)
    studies = q.order_by(Study.created_at.desc()).all()
    return render_template('cms/admin/studies_list.html', studies=studies,
                           show_archived=show_archived)


@admin_bp.route('/studies/new', methods=['GET', 'POST'])
@admin_required
def study_create():
    from app.models.study import Study
    from app.models.survey import Survey
    from app.models.agent import AIAgent
    from app.models.research import Research
    surveys = Survey.query.filter_by(is_active=True).order_by(Survey.name).all()
    tasks = Task.query.filter_by(is_active=True).order_by(Task.sort_order).all()
    agents = AIAgent.query.filter_by(is_active=True).order_by(AIAgent.sort_order).all()
    researches = Research.query.order_by(Research.title).all()

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        if not title:
            flash('Titel ist erforderlich.', 'danger')
            return render_template('cms/admin/study_edit.html', study=None, surveys=surveys,
                                   tasks=tasks, conditions=[], agents=agents, tracking_cfg={})

        _esid = request.form.get('enrollment_survey_id', '') or None
        study = Study(
            title=title,
            description=request.form.get('description', '').strip() or None,
            is_active=bool(request.form.get('is_active')),
            allow_self_enrollment=bool(request.form.get('allow_self_enrollment')),
            one_time_only=bool(request.form.get('one_time_only', True)),
            require_consent=bool(request.form.get('require_consent')),
            consent_text=request.form.get('consent_text', '').strip() or None,
            anonymize_export=bool(request.form.get('anonymize_export', True)),
            study_design=request.form.get('study_design', 'within'),
            is_template=bool(request.form.get('is_template')),
            enrollment_survey_id=int(_esid) if _esid else None,
            leaderboard_enabled=bool(request.form.get('leaderboard_enabled')),
            agent_display_name=request.form.get('agent_display_name', '').strip() or None,
            tracking_config=_build_tracking_config(),
            created_by_id=current_user.id,
        )
        _set_study_dates(study)
        db.session.add(study)
        db.session.flush()
        _save_study_steps(study)
        db.session.commit()
        _rid = request.form.get('research_id', '') or None
        study.research_id = int(_rid) if _rid else None
        db.session.commit()
        log_action('create_study', 'Study', study.id, {'title': title})
        flash(f'Studie "{title}" erstellt.', 'success')
        return redirect(url_for('admin.study_edit', study_id=study.id))

    return render_template('cms/admin/study_edit.html', study=None, surveys=surveys,
                           tasks=tasks, conditions=[], agents=agents, tracking_cfg={}, researches=researches)


@admin_bp.route('/studies/<int:study_id>/edit', methods=['GET', 'POST'])
@admin_required
def study_edit(study_id: int):
    from app.models.study import Study
    from app.models.survey import Survey
    from app.models.agent import AIAgent
    from app.models.research import Research
    study = Study.query.get_or_404(study_id)
    surveys = Survey.query.filter_by(is_active=True).order_by(Survey.name).all()
    tasks = Task.query.filter_by(is_active=True).order_by(Task.sort_order).all()
    agents = AIAgent.query.filter_by(is_active=True).order_by(AIAgent.sort_order).all()
    researches = Research.query.order_by(Research.title).all()

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        if not title:
            flash('Titel ist erforderlich.', 'danger')
            _tc = {}
            if study.tracking_config:
                try:
                    _tc = json.loads(study.tracking_config)
                except Exception:
                    _tc = {}
            return render_template('cms/admin/study_edit.html', study=study, surveys=surveys,
                                   tasks=tasks, conditions=study.conditions, agents=agents,
                                   tracking_cfg=_tc)

        study.title = title
        study.description = request.form.get('description', '').strip() or None
        study.is_active = bool(request.form.get('is_active'))
        study.allow_self_enrollment = bool(request.form.get('allow_self_enrollment'))
        study.one_time_only = bool(request.form.get('one_time_only'))
        study.require_consent = bool(request.form.get('require_consent'))
        study.consent_text = request.form.get('consent_text', '').strip() or None
        study.anonymize_export = bool(request.form.get('anonymize_export'))
        study.study_design = request.form.get('study_design', 'within')
        study.is_template = bool(request.form.get('is_template'))
        _esid = request.form.get('enrollment_survey_id', '') or None
        study.enrollment_survey_id = int(_esid) if _esid else None
        study.leaderboard_enabled = bool(request.form.get('leaderboard_enabled'))
        study.agent_display_name = request.form.get('agent_display_name', '').strip() or None
        study.tracking_config = _build_tracking_config()
        _rid = request.form.get('research_id', '') or None
        study.research_id = int(_rid) if _rid else None
        _set_study_dates(study)
        _save_study_steps(study)
        db.session.commit()
        log_action('edit_study', 'Study', study_id, {'title': title})
        flash(f'Studie "{title}" gespeichert.', 'success')
        return redirect(url_for('admin.study_edit', study_id=study_id))

    _tc = {}
    if study.tracking_config:
        try:
            _tc = json.loads(study.tracking_config)
        except Exception:
            _tc = {}
    return render_template('cms/admin/study_edit.html', study=study, surveys=surveys,
                           tasks=tasks, conditions=study.conditions, agents=agents,
                           tracking_cfg=_tc, researches=researches)


@admin_bp.route('/studies/<int:study_id>/delete', methods=['POST'])
@admin_required
def study_delete(study_id: int):
    from app.models.study import Study
    study = Study.query.get_or_404(study_id)
    title = study.title
    if request.form.get('force') == '1':
        db.session.delete(study)
        db.session.commit()
        log_action('delete_study', 'Study', study_id, {'title': title})
        flash(f'Studie "{title}" unwiderruflich gelöscht.', 'success')
    else:
        study.is_archived = True
        study.archived_at = datetime.now(timezone.utc)
        study.is_active = False
        db.session.commit()
        log_action('archive_study', 'Study', study_id, {'title': title})
        flash(f'Studie "{title}" archiviert (Daten erhalten).', 'info')
    return redirect(url_for('admin.studies_list'))


@admin_bp.route('/studies/<int:study_id>/archive', methods=['POST'])
@admin_required
def study_archive(study_id: int):
    from app.models.study import Study
    study = Study.query.get_or_404(study_id)
    study.is_archived = True
    study.archived_at = datetime.now(timezone.utc)
    study.is_active = False
    db.session.commit()
    log_action('archive_study', 'Study', study_id, {'title': study.title})
    flash(f'Studie "{study.title}" eingefroren.', 'info')
    return redirect(url_for('admin.studies_list'))


@admin_bp.route('/studies/<int:study_id>/unarchive', methods=['POST'])
@admin_required
def study_unarchive(study_id: int):
    from app.models.study import Study
    study = Study.query.get_or_404(study_id)
    study.is_archived = False
    study.archived_at = None
    db.session.commit()
    log_action('unarchive_study', 'Study', study_id, {'title': study.title})
    flash(f'Studie "{study.title}" reaktiviert.', 'success')
    return redirect(url_for('admin.studies_list'))


@admin_bp.route('/studies/<int:study_id>/clone', methods=['POST'])
@admin_required
def study_clone(study_id: int):
    from app.models.study import Study, StudyStep, StudyCondition
    source = Study.query.get_or_404(study_id)
    clone = Study(
        title=source.title + ' (Kopie)',
        description=source.description,
        is_active=False,
        is_template=source.is_template,
        cloned_from_id=source.id,
        allow_self_enrollment=source.allow_self_enrollment,
        max_participants=source.max_participants,
        one_time_only=source.one_time_only,
        require_consent=source.require_consent,
        consent_text=source.consent_text,
        anonymize_export=source.anonymize_export,
        study_design=source.study_design,
        created_by_id=current_user.id,
    )
    db.session.add(clone)
    db.session.flush()
    cond_map = {}
    for cond in source.conditions:
        new_cond = StudyCondition(
            study_id=clone.id,
            name=cond.name,
            description=cond.description,
            target_size=cond.target_size,
        )
        db.session.add(new_cond)
        db.session.flush()
        cond_map[cond.id] = new_cond.id
    for step in source.steps:
        new_step = StudyStep(
            study_id=clone.id,
            step_order=step.step_order,
            step_type=step.step_type,
            survey_id=step.survey_id,
            task_id=step.task_id,
            label=step.label,
            wave_number=step.wave_number,
            available_from=step.available_from,
            available_until=step.available_until,
            allow_late_submission=step.allow_late_submission,
            late_penalty_note=step.late_penalty_note,
            condition_id=cond_map.get(step.condition_id) if step.condition_id else None,
            available_agents=step.available_agents,
            agent_choice_intro=step.agent_choice_intro,
        )
        db.session.add(new_step)
    db.session.commit()
    log_action('clone_study', 'Study', clone.id, {'source_id': source.id})
    flash(f'Studie geklont als "{clone.title}".', 'success')
    return redirect(url_for('admin.study_edit', study_id=clone.id))


# ── Participants ──────────────────────────────────────────────────────────────

@admin_bp.route('/studies/<int:study_id>/participants')
@admin_required
def study_participants(study_id: int):
    from app.models.study import Study, StudyParticipant
    study = Study.query.get_or_404(study_id)
    participants = (StudyParticipant.query
                    .filter_by(study_id=study_id)
                    .order_by(StudyParticipant.enrolled_at.desc())
                    .all())
    all_users = User.query.order_by(User.username).all()
    return render_template('cms/admin/study_participants.html',
                           study=study, participants=participants, all_users=all_users)


@admin_bp.route('/studies/<int:study_id>/participants/add', methods=['POST'])
@admin_required
def study_participant_add(study_id: int):
    from app.models.study import Study, StudyParticipant
    study = Study.query.get_or_404(study_id)
    user_id = request.form.get('user_id', type=int)
    if not user_id:
        flash('Benutzer auswählen.', 'danger')
        return redirect(url_for('admin.study_participants', study_id=study_id))
    existing = StudyParticipant.query.filter_by(study_id=study_id, user_id=user_id).first()
    if existing:
        flash('Benutzer ist bereits angemeldet.', 'warning')
        return redirect(url_for('admin.study_participants', study_id=study_id))
    p = StudyParticipant(study_id=study_id, user_id=user_id)
    db.session.add(p)
    db.session.commit()
    log_action('add_study_participant', 'Study', study_id, {'user_id': user_id})
    flash('Teilnehmer hinzugefügt.', 'success')
    return redirect(url_for('admin.study_participants', study_id=study_id))


@admin_bp.route('/studies/<int:study_id>/participants/<int:participant_id>/remove', methods=['POST'])
@admin_required
def study_participant_remove(study_id: int, participant_id: int):
    from app.models.study import StudyParticipant
    p = StudyParticipant.query.filter_by(id=participant_id, study_id=study_id).first_or_404()
    db.session.delete(p)
    db.session.commit()
    log_action('remove_study_participant', 'Study', study_id, {'participant_id': participant_id})
    flash('Teilnehmer entfernt.', 'success')
    return redirect(url_for('admin.study_participants', study_id=study_id))


@admin_bp.route('/studies/<int:study_id>/participants/<int:participant_id>/reset', methods=['POST'])
@admin_required
def study_participant_reset(study_id: int, participant_id: int):
    from app.models.study import Study, StudyParticipant, StudyStepCompletion, AgentSwitchHistory
    from app.models.task import TaskSubmission
    from app.models.survey import SurveyResponse
    p = StudyParticipant.query.filter_by(id=participant_id, study_id=study_id).first_or_404()
    study = Study.query.get_or_404(study_id)

    # Reset participant state
    p.current_step = 0
    p.completed_at = None
    p.dropped_out_at = None
    p.dropout_reason = None
    p.active_agent_id = None

    # Delete step completions
    StudyStepCompletion.query.filter_by(participant_id=p.id).delete(synchronize_session=False)

    # Delete agent switch history for this participant
    AgentSwitchHistory.query.filter_by(participant_id=p.id).delete(synchronize_session=False)

    # Delete task submissions belonging to this study for this user
    if p.user_id:
        TaskSubmission.query.filter_by(
            study_id=study_id, user_id=p.user_id
        ).delete(synchronize_session=False)

        # Delete survey responses for steps of this study for this user
        step_ids = [s.id for s in study.steps]
        if step_ids:
            SurveyResponse.query.filter(
                SurveyResponse.user_id == p.user_id,
                SurveyResponse.step_id.in_(step_ids)
            ).delete(synchronize_session=False)

    db.session.commit()
    log_action('reset_study_participant', 'Study', study_id,
               {'participant_id': participant_id, 'user_id': p.user_id})
    flash('Studienfortschritt vollständig zurückgesetzt.', 'success')
    return redirect(url_for('admin.study_participants', study_id=study_id))


@admin_bp.route('/studies/<int:study_id>/participants/<int:participant_id>/condition', methods=['POST'])
@admin_required
def study_participant_set_condition(study_id: int, participant_id: int):
    from app.models.study import StudyParticipant
    p = StudyParticipant.query.filter_by(id=participant_id, study_id=study_id).first_or_404()
    raw = request.form.get('condition_id', '') or None
    p.condition_id = int(raw) if raw else None
    db.session.commit()
    log_action('set_participant_condition', 'Study', study_id,
               {'participant_id': participant_id, 'condition_id': p.condition_id})
    flash('Bedingung aktualisiert.', 'success')
    return redirect(url_for('admin.study_participants', study_id=study_id))


@admin_bp.route('/studies/<int:study_id>/broadcast', methods=['POST'])
@admin_required
def study_broadcast(study_id: int):
    from app.models.study import Study, StudyParticipant
    study = Study.query.get_or_404(study_id)
    title = request.form.get('title', '').strip()
    message = request.form.get('message', '').strip()
    condition_id_raw = request.form.get('condition_id', '') or None
    send_email = bool(request.form.get('send_email'))

    if not title:
        flash('Titel ist erforderlich.', 'danger')
        return redirect(url_for('admin.study_participants', study_id=study_id))

    query = StudyParticipant.query.filter_by(study_id=study_id)
    if condition_id_raw:
        query = query.filter_by(condition_id=int(condition_id_raw))
    participants = query.all()

    notifs = []
    for p in participants:
        if p.user_id:
            notifs.append(Notification(
                user_id=p.user_id,
                title=title,
                message=message,
                notif_type='system',
                link=None,
            ))
    if notifs:
        db.session.bulk_save_objects(notifs)
        db.session.commit()

    if send_email:
        for p in participants:
            if p.user and p.user.email:
                try:
                    from app.utils.email import send_study_announcement
                    send_study_announcement(p.user, study, title, message)
                except Exception as exc:
                    current_app.logger.warning('Broadcast email failed for user %s: %s', p.user_id, exc)

    log_action('study_broadcast', 'Study', study_id,
               {'title': title, 'recipients': len(notifs), 'condition_id': condition_id_raw})
    flash(f'{len(notifs)} Nachricht(en) gesendet.', 'success')
    return redirect(url_for('admin.study_participants', study_id=study_id))


# ── Study analytics ───────────────────────────────────────────────────────────

@admin_bp.route('/studies/<int:study_id>/analytics')
@admin_required
def study_analytics(study_id: int):
    from app.models.study import Study, StudyParticipant, StudyStepCompletion
    study = Study.query.get_or_404(study_id)
    participants = StudyParticipant.query.filter_by(study_id=study_id).all()
    total = len(participants)
    completed = sum(1 for p in participants if p.is_completed)
    active = sum(1 for p in participants if p.is_active)
    dropped = sum(1 for p in participants if p.is_dropped_out)
    consent_given = sum(1 for p in participants if p.consent_given_at)

    step_stats = []
    for step in study.steps:
        completions = StudyStepCompletion.query.filter_by(step_id=step.id).all()
        done = sum(1 for c in completions if c.completed_at)
        times = [c.time_spent_seconds for c in completions
                 if c.completed_at and c.time_spent_seconds]
        avg_sec = sum(times) // len(times) if times else None
        late_count = sum(1 for c in completions if c.is_late)
        step_stats.append({
            'step': step,
            'started': len(completions),
            'completed': done,
            'pct': round(100 * done / total) if total else 0,
            'avg_seconds': avg_sec,
            'late': late_count,
        })

    condition_stats = []
    for cond in study.conditions:
        cond_participants = [p for p in participants if p.condition_id == cond.id]
        condition_stats.append({
            'condition': cond,
            'count': len(cond_participants),
            'completed': sum(1 for p in cond_participants if p.is_completed),
        })

    from app.models.task import TaskSubmission as _TaskSub
    task_steps = [s for s in study.steps if s.step_type == 'task' and s.task_id]
    curve_labels = [s.display_label for s in task_steps]

    # Pre-load all submissions for this study in a single query (avoids N+1)
    _task_ids = [s.task_id for s in task_steps]
    _all_subs = (
        _TaskSub.query
        .filter(_TaskSub.study_id == study.id, _TaskSub.task_id.in_(_task_ids))
        .order_by(_TaskSub.id)
        .all()
    )
    # Build lookup: (user_id, task_id) → latest submission (highest id wins, ordered asc so later overwrites)
    _sub_lookup: dict = {}
    for _s in _all_subs:
        _sub_lookup[(_s.user_id, _s.task_id)] = _s

    participant_curves = {}
    group_avg_scores: dict = {}
    for p in participants:
        ident = f'P{p.id:04d}' if study.anonymize_export else (
            p.user.username if p.user else str(p.user_id))
        scores = []
        for step in task_steps:
            sub = _sub_lookup.get((p.user_id, step.task_id))
            val = None
            if sub and sub.grade_value is not None:
                val = round(float(sub.grade_value), 2)
            elif sub and sub.grade_passed is not None:
                val = 1 if sub.grade_passed else 0
            scores.append(val)
            if val is not None:
                group_avg_scores.setdefault(step.step_order, []).append(val)
        participant_curves[ident] = scores
    group_avg = [
        round(sum(group_avg_scores[s.step_order]) / len(group_avg_scores[s.step_order]), 2)
        if s.step_order in group_avg_scores else None
        for s in task_steps
    ]

    curve_data = json.dumps({
        'labels': curve_labels,
        'participants': participant_curves,
        'group_avg': group_avg,
    })

    return render_template('cms/admin/study_analytics.html',
                           study=study,
                           total=total, completed=completed, active=active,
                           dropped=dropped, consent_given=consent_given,
                           step_stats=step_stats,
                           condition_stats=condition_stats,
                           curve_data=curve_data)


# ── Study export ──────────────────────────────────────────────────────────────

@admin_bp.route('/studies/<int:study_id>/export')
@admin_required
def study_export(study_id: int):
    from app.models.study import Study
    study = Study.query.get_or_404(study_id)
    return render_template('cms/admin/study_export.html', study=study)


@admin_bp.route('/studies/<int:study_id>/export/zip')
@admin_required
def study_export_zip(study_id: int):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from app.models.study import Study, StudyParticipant, StudyStepCompletion
    from app.models.survey import SurveyResponse, SurveyQuestion, SurveyPage

    study = Study.query.get_or_404(study_id)
    participants = StudyParticipant.query.filter_by(study_id=study_id).order_by(StudyParticipant.id).all()
    anon = study.anonymize_export

    def _ident(p):
        return f'P{p.id:04d}' if anon else (p.user.username if p.user else str(p.user_id))

    def _sanitize(val) -> str:
        if val is None:
            return ''
        return str(val).replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ')

    def _ws_header(ws, headers):
        """Write a bold header row with light-blue fill to a worksheet."""
        ws.append(headers)
        hdr_fill = PatternFill('solid', fgColor='BDD7EE')
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold
            cell.fill = hdr_fill
            cell.alignment = Alignment(wrap_text=False)

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ── Sheet 1: Participants ──────────────────────────────────────────────────
    ws_p = wb.create_sheet('Participants')
    p_header = ['participant_id', 'user_identifier']
    if not anon:
        p_header += ['user_id', 'email']
    p_header += ['condition', 'enrolled_at', 'completed_at', 'dropped_out',
                 'dropout_reason', 'consent_given_at', 'current_step', 'total_steps', 'progress_pct']
    _ws_header(ws_p, p_header)
    for p in participants:
        row = [p.id, _ident(p)]
        if not anon:
            row += [p.user_id, p.user.email if p.user else '']
        row += [
            p.condition.name if p.condition else '',
            p.enrolled_at.isoformat() if p.enrolled_at else '',
            p.completed_at.isoformat() if p.completed_at else '',
            'yes' if p.is_dropped_out else 'no',
            p.dropout_reason or '',
            p.consent_given_at.isoformat() if p.consent_given_at else '',
            p.current_step,
            study.get_step_count_for_participant(p),
            p.progress_pct,
        ]
        ws_p.append(row)

    # ── Sheet 2: Enrollment Survey ──────────────────────────────────────────────
    # Responses for the pre-study enrollment survey (study.enrollment_survey_id)
    from app.models.survey import Survey as SurveyModel
    if study.enrollment_survey_id:
        enroll_srv = db.session.get(SurveyModel, study.enrollment_survey_id)
        if enroll_srv:
            ws_es = wb.create_sheet('Enrollment_Survey')
            # Build question list
            enroll_qs = []
            for page in enroll_srv.pages:
                for q in page.questions:
                    if q.question_type != 'info':
                        enroll_qs.append(q)
            es_header = ['participant_id', 'user_identifier', 'condition',
                         'response_id', 'completed_at']
            for q in enroll_qs:
                es_header.append(f'q{q.id}_{q.label[:40].replace(chr(10), " ")}')
            _ws_header(ws_es, es_header)
            for p in participants:
                # Find the enrollment survey response submitted after enrollment
                resp = SurveyResponse.query.filter_by(
                    survey_id=study.enrollment_survey_id,
                    user_id=p.user_id,
                ).filter(SurveyResponse.completed_at.isnot(None)).order_by(
                    SurveyResponse.id.desc()
                ).first()
                row = [p.id, _ident(p), p.condition.name if p.condition else '',
                       resp.id if resp else '',
                       resp.completed_at.isoformat() if (resp and resp.completed_at) else '']
                answers = resp.answers if resp else {}
                for q in enroll_qs:
                    row.append(answers.get(str(q.id), ''))
                ws_es.append(row)

    # ── Sheet 3: Survey Responses (step-based) ─────────────────────────────────
    step_surveys = []
    all_questions = []
    for step in study.steps:
        if step.step_type != 'survey' or not step.survey_id:
            continue
        srv = db.session.get(SurveyModel, step.survey_id)
        if not srv:
            continue
        safe_name = srv.name[:20].replace(',', ';').replace('\n', ' ')
        step_surveys.append((step.id, srv.id, safe_name, step.step_order + 1))
        for page in srv.pages:
            for q in page.questions:
                if q.question_type == 'info':
                    continue
                all_questions.append((step.id, srv.id, safe_name, step.step_order + 1,
                                      q.id, q.label[:60].replace('\n', ' ')))

    ws_sr = wb.create_sheet('Survey_Responses')
    sr_header = ['participant_id', 'user_identifier', 'condition']
    q_cols = []
    emitted_step_meta = set()
    for step_id, srv_id, srv_name, step_num, q_id, q_label in all_questions:
        if (step_id, srv_id) not in emitted_step_meta:
            sr_header.append(f's{step_num}_{srv_name}_response_id')
            sr_header.append(f's{step_num}_{srv_name}_completed_at')
            emitted_step_meta.add((step_id, srv_id))
        col_name = f's{step_num}_{srv_name}_{q_id}_{q_label[:30]}'.replace(',', ';').replace('\n', ' ')
        sr_header.append(col_name)
        q_cols.append((step_id, srv_id, q_id))
    _ws_header(ws_sr, sr_header)
    for p in participants:
        row = [p.id, _ident(p), p.condition.name if p.condition else '']
        # Pass 1: exact step_id match
        responses_by_step = {}
        for step_id, srv_id, _, _ in step_surveys:
            resp = SurveyResponse.query.filter_by(
                survey_id=srv_id, user_id=p.user_id, step_id=step_id
            ).order_by(SurveyResponse.id.desc()).first()
            responses_by_step[(step_id, srv_id)] = resp
        # Pass 2: fallback for responses where step_id was cleared (ondelete='SET NULL')
        # or was never stored (pre-migration data). Assign NULL-step_id responses to
        # unmatched steps in chronological order so the same survey used multiple times
        # maps oldest response → earliest step, next → next step, etc.
        null_pool: dict = {}
        for step_id, srv_id, _, _ in step_surveys:
            if responses_by_step[(step_id, srv_id)] is not None:
                continue
            if srv_id not in null_pool:
                null_pool[srv_id] = (
                    SurveyResponse.query
                    .filter_by(survey_id=srv_id, user_id=p.user_id)
                    .filter(SurveyResponse.step_id.is_(None))
                    .filter(SurveyResponse.completed_at.isnot(None))
                    .order_by(SurveyResponse.id.asc())
                    .all()
                )
            pool = null_pool[srv_id]
            if pool:
                responses_by_step[(step_id, srv_id)] = pool.pop(0)
        # Pass 3: single-response fallback (catches any remaining edge cases)
        for step_id, srv_id, _, _ in step_surveys:
            if responses_by_step[(step_id, srv_id)] is not None:
                continue
            all_resps = SurveyResponse.query.filter_by(
                survey_id=srv_id, user_id=p.user_id
            ).all()
            if len(all_resps) == 1:
                responses_by_step[(step_id, srv_id)] = all_resps[0]
        emitted_meta = set()
        for step_id, srv_id, q_id in q_cols:
            step_key = (step_id, srv_id)
            if step_key not in emitted_meta:
                resp = responses_by_step.get(step_key)
                row.append(resp.id if resp else '')
                row.append(resp.completed_at.isoformat() if (resp and resp.completed_at) else '')
                emitted_meta.add(step_key)
            answers = responses_by_step.get(step_key)
            row.append((answers.answers if answers else {}).get(str(q_id), ''))
        ws_sr.append(row)

    # ── Sheet 4: Task Submissions ──────────────────────────────────────────────
    ws_sub = wb.create_sheet('Task_Submissions')
    sub_header = ['participant_id', 'user_identifier', 'condition',
                  'task_id', 'task_title', 'step_number', 'wave_number',
                  'agent_id', 'agent_name',
                  'started_at', 'completed_at', 'time_spent_seconds',
                  'interactions', 'tokens_in', 'tokens_out',
                  'grade_value', 'grade_passed', 'grade_comment', 'submission_id']
    _ws_header(ws_sub, sub_header)
    task_step_map = {s.task_id: s for s in study.steps if s.step_type == 'task' and s.task_id}
    for p in participants:
        subs = TaskSubmission.query.filter_by(study_id=study_id, user_id=p.user_id).all()
        for sub in subs:
            step = task_step_map.get(sub.task_id)
            duration = ''
            if sub.started_at and sub.completed_at:
                duration = int((sub.completed_at - sub.started_at).total_seconds())
            ws_sub.append([
                p.id, _ident(p), p.condition.name if p.condition else '',
                sub.task_id,
                sub.task.title if sub.task else sub.task_id,
                (step.step_order + 1) if step else '',
                step.wave_number if step else '',
                sub.agent_id or '',
                sub.agent_name or '',
                sub.started_at.isoformat() if sub.started_at else '',
                sub.completed_at.isoformat() if sub.completed_at else '',
                duration,
                sub.interactions or 0,
                sub.tokens_in or 0,
                sub.tokens_out or 0,
                sub.grade_value if sub.grade_value is not None else '',
                'pass' if sub.grade_passed else ('fail' if sub.grade_passed is False else ''),
                _sanitize(sub.grade_comment or ''),
                sub.id,
            ])

    # ── Sheet 5: Step Timings ──────────────────────────────────────────────────
    ws_t = wb.create_sheet('Step_Timings')
    t_header = ['participant_id', 'user_identifier',
                'step_number', 'step_label', 'step_type', 'wave_number',
                'condition', 'started_at', 'completed_at',
                'time_spent_seconds', 'is_late']
    _ws_header(ws_t, t_header)
    for p in participants:
        for step in study.steps:
            comp = StudyStepCompletion.query.filter_by(
                participant_id=p.id, step_id=step.id).first()
            if comp:
                ws_t.append([
                    p.id, _ident(p),
                    step.step_order + 1, step.display_label, step.step_type, step.wave_number,
                    p.condition.name if p.condition else '',
                    comp.started_at.isoformat() if comp.started_at else '',
                    comp.completed_at.isoformat() if comp.completed_at else '',
                    comp.time_spent_seconds or '',
                    'yes' if comp.is_late else 'no',
                ])

    # ── Sheet 6: Tracking Events (raw batches) ────────────────────────────────
    from app.models.tracking import TaskSessionTracking
    ws_tr = wb.create_sheet('Tracking_Events')
    tr_header = ['participant_id', 'user_identifier', 'condition',
                 'task_id', 'submission_id', 'session_start',
                 'batch_seq', 'event_count', 'events_json']
    _ws_header(ws_tr, tr_header)
    for p in participants:
        trackings = (TaskSessionTracking.query
                     .filter_by(study_id=study_id, user_id=p.user_id)
                     .order_by(TaskSessionTracking.task_id, TaskSessionTracking.batch_seq)
                     .all())
        for tr in trackings:
            try:
                evts = json.loads(tr.events_data) if tr.events_data else []
                ev_count = len(evts) if isinstance(evts, list) else 0
            except Exception:
                ev_count = 0
            ws_tr.append([
                p.id, _ident(p), p.condition.name if p.condition else '',
                tr.task_id, tr.submission_id or '',
                str(tr.session_start or ''),
                tr.batch_seq,
                ev_count,
                tr.events_data or '[]',
            ])

    # ── Sheet 7: BPMN Element Events (flattened) ──────────────────────────────
    # One row per BPMN diagram change (add/remove/connect/rename/move).
    # Combines both user and AI actions; 'source' column distinguishes them.
    _BPMN_ETYPES = {'bpmn_add', 'bpmn_remove', 'bpmn_connect',
                    'bpmn_disconnect', 'bpmn_rename', 'bpmn_move'}
    ws_bpmn = wb.create_sheet('BPMN_Element_Events')
    _ws_header(ws_bpmn, [
        'participant_id', 'user_identifier', 'condition', 'task_id',
        'timestamp', 'event_type', 'source',
        'element_id', 'element_type', 'element_name',
        'source_element_id', 'source_element_name',
        'target_element_id', 'target_element_name',
        'x', 'y',
    ])
    for p in participants:
        trackings = (TaskSessionTracking.query
                     .filter_by(study_id=study_id, user_id=p.user_id)
                     .order_by(TaskSessionTracking.task_id, TaskSessionTracking.batch_seq)
                     .all())
        for tr in trackings:
            try:
                evts = json.loads(tr.events_data) if tr.events_data else []
            except Exception:
                evts = []
            for ev in evts:
                if not isinstance(ev, dict):
                    continue
                if ev.get('type') not in _BPMN_ETYPES:
                    continue
                ws_bpmn.append([
                    p.id,
                    _ident(p),
                    p.condition.name if p.condition else '',
                    tr.task_id,
                    ev.get('ts', ''),
                    ev.get('type', ''),
                    ev.get('source', ''),
                    ev.get('element_id', ''),
                    ev.get('element_type', ''),
                    ev.get('element_name', ''),
                    ev.get('source_element_id', ''),
                    ev.get('source_element_name', ''),
                    ev.get('target_element_id', ''),
                    ev.get('target_element_name', ''),
                    ev.get('x', ''),
                    ev.get('y', ''),
                ])

    # ── Sheet 8: LLM Interactions ──────────────────────────────────────────────
    ws_llm = wb.create_sheet('LLM_Interactions')
    llm_header = ['participant_id', 'user_identifier', 'condition',
                  'task_id', 'submission_id', 'submission_started_at',
                  'call_index', 'timestamp', 'phase_label',
                  'message_count', 'system_prompt', 'full_messages_json', 'response']
    _ws_header(ws_llm, llm_header)
    for p in participants:
        subs = TaskSubmission.query.filter_by(study_id=study_id, user_id=p.user_id).all()
        for sub in subs:
            if not sub.llm_prompt_log:
                continue
            try:
                calls = json.loads(sub.llm_prompt_log)
            except Exception:
                continue
            if not isinstance(calls, list):
                continue
            for idx, call in enumerate(calls):
                messages = call.get('messages', [])
                system_prompt = ''
                for m in messages:
                    if m.get('role') == 'system':
                        system_prompt = m.get('content', '')
                        break
                ws_llm.append([
                    p.id, _ident(p), p.condition.name if p.condition else '',
                    sub.task_id, sub.id,
                    sub.started_at.isoformat() if sub.started_at else '',
                    idx + 1,
                    call.get('ts', ''),
                    call.get('label', ''),
                    len(messages),
                    _sanitize(system_prompt),
                    _sanitize(json.dumps(messages, ensure_ascii=False)),
                    _sanitize(call.get('response', '')),
                ])

    # ── Sheet 8: Agent Choices ─────────────────────────────────────────────────
    from app.models.study import AgentSwitchHistory
    ws_ac = wb.create_sheet('Agent_Choices')
    ac_header = ['participant_id', 'user_identifier', 'condition',
                 'wave_number', 'step_label', 'chosen_agent_id', 'chosen_agent_name',
                 'previous_agent_id', 'previous_agent_name', 'chosen_at']
    _ws_header(ws_ac, ac_header)
    for p in participants:
        switches = (AgentSwitchHistory.query
                    .join(AgentSwitchHistory.participant)
                    .filter(AgentSwitchHistory.participant_id == p.id)
                    .order_by(AgentSwitchHistory.chosen_at)
                    .all())
        for sw in switches:
            step_lbl = sw.step.display_label if sw.step else ''
            ws_ac.append([
                p.id, _ident(p), p.condition.name if p.condition else '',
                sw.wave_number, step_lbl,
                sw.agent_id or '',
                sw.agent.name if sw.agent else '',
                sw.previous_agent_id or '',
                sw.previous_agent.name if sw.previous_agent else '',
                sw.chosen_at.isoformat() if sw.chosen_at else '',
            ])

    # ── Build ZIP with Excel + BPMN files ─────────────────────────────────────
    wb_buf = io.BytesIO()
    wb.save(wb_buf)
    wb_buf.seek(0)

    ts = datetime.now().strftime('%Y%m%d_%H%M')
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f'study_{study_id}_data.xlsx', wb_buf.read())
        for p in participants:
            subs = TaskSubmission.query.filter_by(study_id=study_id, user_id=p.user_id).all()
            for sub in subs:
                if sub.bpmn_xml:
                    fname = f'bpmn/{_ident(p)}_{sub.task_id}_sub{sub.id}.xml'
                    zf.writestr(fname, sub.bpmn_xml.encode('utf-8'))

    zip_buf.seek(0)
    return Response(
        zip_buf.read(),
        mimetype='application/zip',
        headers={
            'Content-Disposition': f'attachment; filename="study_{study_id}_export_{ts}.zip"'
        }
    )


@admin_bp.route('/studies/<int:study_id>/export/participants.csv')
@admin_required
def study_export_participants_csv(study_id: int):
    from app.models.study import Study, StudyParticipant, StudyStepCompletion
    study = Study.query.get_or_404(study_id)
    participants = StudyParticipant.query.filter_by(study_id=study_id).order_by(StudyParticipant.id).all()
    anon = study.anonymize_export

    buf = io.StringIO()
    w = csv.writer(buf)
    header = ['participant_id', 'user_identifier']
    if not anon:
        header += ['user_id', 'email']
    header += ['condition', 'enrolled_at', 'completed_at', 'dropped_out',
               'consent_given', 'current_step', 'progress_pct']
    for step in study.steps:
        lbl = step.display_label[:30].replace(',', ';')
        header += [f's{step.step_order+1}_{lbl}_started',
                   f's{step.step_order+1}_{lbl}_completed',
                   f's{step.step_order+1}_{lbl}_seconds',
                   f's{step.step_order+1}_{lbl}_late']
    w.writerow(header)

    for p in participants:
        ident = f'P{p.id:04d}' if anon else (p.user.username if p.user else '')
        row = [p.id, ident]
        if not anon:
            row += [p.user_id, p.user.email if p.user else '']
        row += [
            p.condition.name if p.condition else '',
            p.enrolled_at.isoformat() if p.enrolled_at else '',
            p.completed_at.isoformat() if p.completed_at else '',
            'yes' if p.is_dropped_out else 'no',
            p.consent_given_at.isoformat() if p.consent_given_at else '',
            p.current_step,
            p.progress_pct,
        ]
        for step in study.steps:
            comp = StudyStepCompletion.query.filter_by(participant_id=p.id, step_id=step.id).first()
            if comp:
                row += [
                    comp.started_at.isoformat() if comp.started_at else '',
                    comp.completed_at.isoformat() if comp.completed_at else '',
                    comp.time_spent_seconds or '',
                    'yes' if comp.is_late else 'no',
                ]
            else:
                row += ['', '', '', '']
        w.writerow(row)

    resp = Response(buf.getvalue().encode('utf-8-sig'), mimetype='text/csv')
    resp.headers['Content-Disposition'] = f'attachment; filename="study_{study_id}_participants.csv"'
    return resp


# ── Conditions ────────────────────────────────────────────────────────────────

@admin_bp.route('/studies/<int:study_id>/conditions', methods=['GET', 'POST'])
@admin_required
def study_conditions(study_id: int):
    from app.models.study import Study, StudyCondition
    study = Study.query.get_or_404(study_id)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        next_url = request.form.get('next') or url_for('admin.study_conditions', study_id=study_id)
        if not name:
            flash('Name ist erforderlich.', 'danger')
        else:
            cond = StudyCondition(
                study_id=study_id,
                name=name,
                description=request.form.get('description', '').strip() or None,
                target_size=int(request.form.get('target_size') or 0) or None,
                agent_id=request.form.get('agent_id') or None,
            )
            db.session.add(cond)
            db.session.commit()
            flash(f'Bedingung "{name}" erstellt.', 'success')
        return redirect(next_url)
    return render_template('cms/admin/study_conditions.html', study=study,
                           conditions=study.conditions)


@admin_bp.route('/studies/<int:study_id>/conditions/<int:cond_id>/delete', methods=['POST'])
@admin_required
def study_condition_delete(study_id: int, cond_id: int):
    from app.models.study import StudyCondition
    cond = StudyCondition.query.filter_by(id=cond_id, study_id=study_id).first_or_404()
    name = cond.name
    next_url = request.form.get('next') or url_for('admin.study_conditions', study_id=study_id)
    db.session.delete(cond)
    db.session.commit()
    flash(f'Bedingung "{name}" gelöscht.', 'success')
    return redirect(next_url)


@admin_bp.route('/studies/<int:study_id>/conditions/<int:cond_id>/update', methods=['POST'])
@admin_required
def study_condition_update(study_id: int, cond_id: int):
    from app.models.study import StudyCondition
    cond = StudyCondition.query.filter_by(id=cond_id, study_id=study_id).first_or_404()
    name = request.form.get('name', '').strip()
    next_url = request.form.get('next') or url_for('admin.study_conditions', study_id=study_id)
    if not name:
        flash('Name ist erforderlich.', 'danger')
        return redirect(next_url)
    cond.name = name
    cond.description = request.form.get('description', '').strip() or None
    raw_ts = request.form.get('target_size', '').strip()
    cond.target_size = int(raw_ts) if raw_ts.isdigit() else None
    agent_id = request.form.get('agent_id', '').strip()
    cond.agent_id = int(agent_id) if agent_id.isdigit() else None
    db.session.commit()
    flash(f'Bedingung "{name}" aktualisiert.', 'success')
    return redirect(next_url)
