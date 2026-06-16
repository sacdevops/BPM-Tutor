"""Admin — Research management routes."""
import io
import json
import zipfile
from datetime import datetime, timezone

from flask import render_template, redirect, url_for, flash, request, Response
from flask_login import current_user

from app.blueprints.admin import admin_bp
from app.extensions import db
from app.models.user import User
from app.utils.decorators import admin_required
from app.utils.audit import log_action


def _parse_dt(key: str):
    val = request.form.get(key, '').strip()
    if not val:
        return None
    try:
        return datetime.fromisoformat(val).replace(tzinfo=timezone.utc)
    except Exception:
        return None


# ── Research CRUD ──────────────────────────────────────────────────────────────

@admin_bp.route('/research')
@admin_required
def research_list():
    """Redirect to the single Research (or to create if none exists)."""
    from app.models.research import Research
    r = Research.query.first()
    if r:
        return redirect(url_for('admin.research_edit', research_id=r.id))
    return redirect(url_for('admin.research_create'))


@admin_bp.route('/research/new', methods=['GET', 'POST'])
@admin_required
def research_create():
    from app.models.research import Research
    from app.models.survey import Survey
    from app.models.agent import AIAgent
    # Enforce singleton — if one already exists, redirect to it
    existing = Research.query.first()
    if existing:
        return redirect(url_for('admin.research_edit', research_id=existing.id))
    surveys = Survey.query.filter_by(is_active=True).order_by(Survey.name).all()
    agents = AIAgent.query.filter_by(is_active=True).order_by(AIAgent.sort_order).all()

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        if not title:
            flash('Titel ist erforderlich.', 'danger')
            return render_template('cms/admin/research_edit.html', research=None,
                                   surveys=surveys, agents=agents, conditions=[])
        _esid = request.form.get('enrollment_survey_id', '') or None
        r = Research(
            title=title,
            description=request.form.get('description', '').strip() or None,
            is_active=bool(request.form.get('is_active')),
            is_enabled=True,
            allow_self_enrollment=bool(request.form.get('allow_self_enrollment')),
            one_time_only=bool(request.form.get('one_time_only', True)),
            require_consent=bool(request.form.get('require_consent')),
            consent_text=request.form.get('consent_text', '').strip() or None,
            anonymize_export=bool(request.form.get('anonymize_export', True)),
            study_design=request.form.get('study_design', 'within'),
            auto_dropout_on_miss=bool(request.form.get('auto_dropout_on_miss')),
            enrollment_survey_id=int(_esid) if _esid else None,
            leaderboard_enabled=bool(request.form.get('leaderboard_enabled')),
            agent_display_name=request.form.get('agent_display_name', '').strip() or None,
            created_by_id=current_user.id,
        )
        r.enrollment_start = _parse_dt('enrollment_start')
        r.enrollment_end = _parse_dt('enrollment_end')
        try:
            max_p = int(request.form.get('max_participants', '') or 0)
            r.max_participants = max_p or None
        except (ValueError, TypeError):
            r.max_participants = None

        db.session.add(r)
        db.session.flush()
        _save_conditions(r)
        db.session.commit()
        log_action('create_research', 'Research', r.id, {'title': title})
        flash(f'Research „{title}" erstellt.', 'success')
        return redirect(url_for('admin.research_edit', research_id=r.id))

    return render_template('cms/admin/research_edit.html', research=None,
                           surveys=surveys, agents=agents, conditions=[])


@admin_bp.route('/research/<int:research_id>/edit', methods=['GET', 'POST'])
@admin_required
def research_edit(research_id: int):
    from app.models.research import Research
    from app.models.survey import Survey
    from app.models.agent import AIAgent
    r = Research.query.get_or_404(research_id)
    surveys = Survey.query.filter_by(is_active=True).order_by(Survey.name).all()
    agents = AIAgent.query.filter_by(is_active=True).order_by(AIAgent.sort_order).all()

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        if not title:
            flash('Titel ist erforderlich.', 'danger')
            return render_template('cms/admin/research_edit.html', research=r,
                                   surveys=surveys, agents=agents, conditions=r.conditions)
        _esid = request.form.get('enrollment_survey_id', '') or None
        r.title = title
        r.description = request.form.get('description', '').strip() or None
        r.is_active = bool(request.form.get('is_active'))
        r.allow_self_enrollment = bool(request.form.get('allow_self_enrollment'))
        r.one_time_only = bool(request.form.get('one_time_only'))
        r.require_consent = bool(request.form.get('require_consent'))
        r.consent_text = request.form.get('consent_text', '').strip() or None
        r.anonymize_export = bool(request.form.get('anonymize_export'))
        r.study_design = request.form.get('study_design', 'within')
        r.auto_dropout_on_miss = bool(request.form.get('auto_dropout_on_miss'))
        r.enrollment_survey_id = int(_esid) if _esid else None
        r.leaderboard_enabled = bool(request.form.get('leaderboard_enabled'))
        r.agent_display_name = request.form.get('agent_display_name', '').strip() or None
        r.enrollment_start = _parse_dt('enrollment_start')
        r.enrollment_end = _parse_dt('enrollment_end')
        try:
            max_p = int(request.form.get('max_participants', '') or 0)
            r.max_participants = max_p or None
        except (ValueError, TypeError):
            r.max_participants = None

        _save_conditions(r)
        db.session.commit()
        log_action('edit_research', 'Research', research_id, {'title': title})
        flash(f'Research „{title}" gespeichert.', 'success')
        return redirect(url_for('admin.research_edit', research_id=research_id))

    return render_template('cms/admin/research_edit.html', research=r,
                           surveys=surveys, agents=agents, conditions=r.conditions)


@admin_bp.route('/research/<int:research_id>/delete', methods=['POST'])
@admin_required
def research_delete(research_id: int):
    from app.models.research import Research
    r = Research.query.get_or_404(research_id)
    title = r.title
    db.session.delete(r)
    db.session.commit()
    log_action('delete_research', 'Research', research_id, {'title': title})
    flash(f'Research „{title}" gelöscht.', 'success')
    return redirect(url_for('admin.research_list'))


def _save_conditions(research) -> None:
    """Rebuild ResearchConditions from POSTed JSON (conditions_json field)."""
    from app.models.research import ResearchCondition
    raw = (request.form.get('conditions_json') or '').strip()
    if not raw:
        # conditions_json is empty — JS did not run (e.g. template error);
        # preserve the existing conditions instead of wiping them.
        return
    try:
        conds_data = json.loads(raw)
    except Exception:
        # Invalid JSON — preserve existing conditions
        return

    ResearchCondition.query.filter_by(research_id=research.id).delete()
    db.session.flush()
    for c in conds_data:
        name = (c.get('name') or '').strip()
        if not name:
            continue
        cond = ResearchCondition(
            research_id=research.id,
            name=name,
            description=(c.get('description') or '').strip() or None,
            target_size=int(c['target_size']) if c.get('target_size') else None,
            agent_id=c.get('agent_id') or None,
        )
        db.session.add(cond)


# ── Research Participants ─────────────────────────────────────────────────────

@admin_bp.route('/research/<int:research_id>/participants')
@admin_required
def research_participants(research_id: int):
    from app.models.research import Research, ResearchParticipant
    r = Research.query.get_or_404(research_id)
    participants = (ResearchParticipant.query
                    .filter_by(research_id=research_id)
                    .order_by(ResearchParticipant.enrolled_at.desc())
                    .all())
    all_users = User.query.order_by(User.username).all()
    return render_template('cms/admin/research_participants.html',
                           research=r, participants=participants, all_users=all_users)


@admin_bp.route('/research/<int:research_id>/participants/add', methods=['POST'])
@admin_required
def research_participant_add(research_id: int):
    from app.models.research import Research, ResearchParticipant
    r = Research.query.get_or_404(research_id)
    user_id = request.form.get('user_id', type=int)
    if not user_id:
        flash('Benutzer auswählen.', 'danger')
        return redirect(url_for('admin.research_participants', research_id=research_id))
    existing = ResearchParticipant.query.filter_by(research_id=research_id, user_id=user_id).first()
    if existing:
        flash('Benutzer ist bereits angemeldet.', 'warning')
        return redirect(url_for('admin.research_participants', research_id=research_id))
    p = ResearchParticipant(research_id=research_id, user_id=user_id)
    r.assign_condition(p)
    if p.condition_id and p.condition and p.condition.agent_id:
        p.active_agent_id = p.condition.agent_id
    db.session.add(p)
    db.session.commit()
    log_action('add_research_participant', 'Research', research_id, {'user_id': user_id})
    flash('Teilnehmer hinzugefügt.', 'success')
    return redirect(url_for('admin.research_participants', research_id=research_id))


@admin_bp.route('/research/<int:research_id>/participants/<int:participant_id>/remove', methods=['POST'])
@admin_required
def research_participant_remove(research_id: int, participant_id: int):
    from app.models.research import Research, ResearchParticipant
    from app.models.study import StudyParticipant
    p = ResearchParticipant.query.filter_by(id=participant_id, research_id=research_id).first_or_404()
    user_id = p.user_id
    # Also remove the user from all sub-studies of this Research
    research = Research.query.get(research_id)
    if research:
        study_ids = [s.id for s in research.studies]
        if study_ids:
            StudyParticipant.query.filter(
                StudyParticipant.user_id == user_id,
                StudyParticipant.study_id.in_(study_ids)
            ).delete(synchronize_session=False)
    db.session.delete(p)
    db.session.commit()
    log_action('remove_research_participant', 'Research', research_id, {'participant_id': participant_id})
    flash('Teilnehmer entfernt.', 'success')
    return redirect(url_for('admin.research_participants', research_id=research_id))


@admin_bp.route('/research/<int:research_id>/participants/<int:participant_id>/dropout', methods=['POST'])
@admin_required
def research_participant_dropout(research_id: int, participant_id: int):
    from app.models.research import ResearchParticipant
    p = ResearchParticipant.query.filter_by(id=participant_id, research_id=research_id).first_or_404()
    if not p.is_dropped_out:
        p.dropped_out_at = datetime.now(timezone.utc)
        p.dropout_reason = request.form.get('reason', 'Manuell durch Admin ausgeschlossen.')[:500]
        db.session.commit()
        flash('Teilnehmer ausgeschlossen.', 'info')
    return redirect(url_for('admin.research_participants', research_id=research_id))


# ── Research Export ───────────────────────────────────────────────────────────

@admin_bp.route('/research/<int:research_id>/export/zip')
@admin_required
def research_export_zip(research_id: int):
    """Export all Research data as a ZIP with Excel workbooks per study."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from app.models.research import Research, ResearchParticipant
    from app.models.study import Study, StudyParticipant, StudyStepCompletion, AgentSwitchHistory
    from app.models.task import TaskSubmission
    from app.models.survey import Survey as SurveyModel, SurveyResponse
    from app.models.tracking import TaskSessionTracking

    research = Research.query.get_or_404(research_id)
    anon = research.anonymize_export

    def _sanitize(val) -> str:
        if val is None:
            return ''
        return str(val).replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ')

    def _slug(text: str, max_len: int = 30) -> str:
        import re
        s = re.sub(r'[^\w\s-]', '', text.lower())
        s = re.sub(r'[\s_-]+', '_', s).strip('_')
        return s[:max_len] or 'study'

    def _ws_header(ws, headers):
        ws.append(headers)
        hdr_fill = PatternFill('solid', fgColor='BDD7EE')
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold
            cell.fill = hdr_fill

    # ── Research-level participants ───────────────────────────────────────────
    rp_all = (ResearchParticipant.query
              .filter_by(research_id=research_id)
              .order_by(ResearchParticipant.id)
              .all())

    def _rident(rp):
        return f'P{rp.id:04d}' if anon else (rp.user.username if rp.user else str(rp.user_id))

    # Build user_id → rident lookup
    rident_map = {rp.user_id: _rident(rp) for rp in rp_all}

    # Top-level workbook with participants + enrollment survey
    top_wb = Workbook()
    top_wb.remove(top_wb.active)

    # Participants sheet
    ws_rp = top_wb.create_sheet('Participants')
    rp_header = ['participant_id', 'user_identifier']
    if not anon:
        rp_header += ['user_id', 'email']
    rp_header += ['condition', 'enrolled_at', 'consent_given_at',
                  'dropped_out', 'dropout_reason']
    _ws_header(ws_rp, rp_header)
    for rp in rp_all:
        row = [rp.id, _rident(rp)]
        if not anon:
            row += [rp.user_id, rp.user.email if rp.user else '']
        row += [
            rp.condition.name if rp.condition else '',
            rp.enrolled_at.isoformat() if rp.enrolled_at else '',
            rp.consent_given_at.isoformat() if rp.consent_given_at else '',
            'yes' if rp.is_dropped_out else 'no',
            rp.dropout_reason or '',
        ]
        ws_rp.append(row)

    # Research enrollment survey (if configured)
    if research.enrollment_survey_id:
        enroll_srv = db.session.get(SurveyModel, research.enrollment_survey_id)
        if enroll_srv:
            ws_es = top_wb.create_sheet('Enrollment_Survey')
            enroll_qs = []
            for page in enroll_srv.pages:
                for q in page.questions:
                    if q.question_type != 'info':
                        enroll_qs.append(q)
            es_header = ['participant_id', 'user_identifier', 'condition',
                         'response_id', 'completed_at']
            for q in enroll_qs:
                es_header.append(f'q{q.id}_{q.label[:40].replace(chr(10), " ")}')
            _ws_header(ws_es, es_header)
            for rp in rp_all:
                resp = SurveyResponse.query.filter_by(
                    survey_id=research.enrollment_survey_id,
                    user_id=rp.user_id,
                ).filter(SurveyResponse.completed_at.isnot(None)).order_by(
                    SurveyResponse.id.desc()
                ).first()
                row = [rp.id, _rident(rp),
                       rp.condition.name if rp.condition else '',
                       resp.id if resp else '',
                       resp.completed_at.isoformat() if (resp and resp.completed_at) else '']
                answers = resp.answers if resp else {}
                for q in enroll_qs:
                    row.append(answers.get(str(q.id), ''))
                ws_es.append(row)

    # ── Per-study data ────────────────────────────────────────────────────────
    zip_buf = io.BytesIO()
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Save top-level workbook
        top_buf = io.BytesIO()
        top_wb.save(top_buf)
        top_buf.seek(0)
        zf.writestr('research_participants.xlsx', top_buf.read())

        for study in sorted(research.studies, key=lambda s: (s.created_at or datetime.min)):
            if study.is_archived:
                continue
            slug = _slug(study.title)
            prefix = f'studies/{slug}/'
            participants = (StudyParticipant.query
                            .filter_by(study_id=study.id)
                            .order_by(StudyParticipant.id)
                            .all())

            def _ident(sp):
                rid = rident_map.get(sp.user_id)
                if rid:
                    return rid
                if anon:
                    return f'P{sp.id:04d}'
                return sp.user.username if sp.user else str(sp.user_id)

            wb = Workbook()
            wb.remove(wb.active)

            # Sheet: Participants / progress
            ws_p = wb.create_sheet('Participants')
            _ws_header(ws_p, ['participant_id', 'study', 'enrolled_at', 'completed_at',
                               'dropped_out', 'current_step', 'progress_pct'])
            for sp in participants:
                ws_p.append([
                    _ident(sp), study.title,
                    sp.enrolled_at.isoformat() if sp.enrolled_at else '',
                    sp.completed_at.isoformat() if sp.completed_at else '',
                    'yes' if sp.is_dropped_out else 'no',
                    sp.current_step, sp.progress_pct,
                ])

            # Sheet: Enrollment Survey (study-level if different from research-level)
            if study.enrollment_survey_id and study.enrollment_survey_id != getattr(research, 'enrollment_survey_id', None):
                srv_enroll = db.session.get(SurveyModel, study.enrollment_survey_id)
                if srv_enroll:
                    ws_se = wb.create_sheet('Study_Enrollment_Survey')
                    se_qs = []
                    for page in srv_enroll.pages:
                        for q in page.questions:
                            if q.question_type != 'info':
                                se_qs.append(q)
                    se_header = ['participant_id', 'response_id', 'completed_at']
                    for q in se_qs:
                        se_header.append(f'q{q.id}_{q.label[:40].replace(chr(10), " ")}')
                    _ws_header(ws_se, se_header)
                    for sp in participants:
                        resp = SurveyResponse.query.filter_by(
                            survey_id=study.enrollment_survey_id,
                            user_id=sp.user_id,
                        ).filter(SurveyResponse.completed_at.isnot(None)).order_by(
                            SurveyResponse.id.desc()
                        ).first()
                        row = [_ident(sp),
                               resp.id if resp else '',
                               resp.completed_at.isoformat() if (resp and resp.completed_at) else '']
                        answers = resp.answers if resp else {}
                        for q in se_qs:
                            row.append(answers.get(str(q.id), ''))
                        ws_se.append(row)

            # Sheet: Survey Responses (step-based)
            step_surveys = []
            all_questions = []
            for step in study.steps:
                if step.step_type != 'survey' or not step.survey_id:
                    continue
                srv = db.session.get(SurveyModel, step.survey_id)
                if not srv:
                    continue
                safe_name = srv.name[:20].replace(',', ';').replace('\n', ' ')
                step_surveys.append((step.id, srv.id, safe_name, step.step_order + 1))
                for page in srv.pages:
                    for q in page.questions:
                        if q.question_type == 'info':
                            continue
                        all_questions.append((step.id, srv.id, safe_name,
                                              step.step_order + 1, q.id,
                                              q.label[:60].replace('\n', ' ')))
            ws_sr = wb.create_sheet('Survey_Responses')
            sr_header = ['participant_id', 'condition']
            q_cols = []
            emitted_step_meta = set()
            for step_id, srv_id, srv_name, step_num, q_id, q_label in all_questions:
                if (step_id, srv_id) not in emitted_step_meta:
                    sr_header.append(f's{step_num}_{srv_name}_response_id')
                    sr_header.append(f's{step_num}_{srv_name}_completed_at')
                    emitted_step_meta.add((step_id, srv_id))
                col_name = f's{step_num}_{srv_name}_{q_id}_{q_label[:30]}'.replace(',', ';').replace('\n', ' ')
                sr_header.append(col_name)
                q_cols.append((step_id, srv_id, q_id))
            _ws_header(ws_sr, sr_header)
            for sp in participants:
                row = [_ident(sp), sp.condition.name if sp.condition else '']
                # Pass 1: exact step_id match
                responses_by_step = {}
                for step_id, srv_id, _, _ in step_surveys:
                    resp = SurveyResponse.query.filter_by(
                        survey_id=srv_id, user_id=sp.user_id, step_id=step_id
                    ).order_by(SurveyResponse.id.desc()).first()
                    responses_by_step[(step_id, srv_id)] = resp
                # Pass 2: fallback for responses where step_id was cleared
                # (ondelete='SET NULL') or was never stored (pre-migration data).
                # Assign NULL-step_id responses in chronological order so the
                # same survey used for multiple steps is assigned correctly.
                null_pool: dict = {}
                for step_id, srv_id, _, _ in step_surveys:
                    if responses_by_step[(step_id, srv_id)] is not None:
                        continue
                    if srv_id not in null_pool:
                        null_pool[srv_id] = (
                            SurveyResponse.query
                            .filter_by(survey_id=srv_id, user_id=sp.user_id)
                            .filter(SurveyResponse.step_id.is_(None))
                            .filter(SurveyResponse.completed_at.isnot(None))
                            .order_by(SurveyResponse.id.asc())
                            .all()
                        )
                    pool = null_pool[srv_id]
                    if pool:
                        responses_by_step[(step_id, srv_id)] = pool.pop(0)
                # Pass 3: single-response fallback (catches remaining edge cases)
                for step_id, srv_id, _, _ in step_surveys:
                    if responses_by_step[(step_id, srv_id)] is not None:
                        continue
                    all_resps = SurveyResponse.query.filter_by(
                        survey_id=srv_id, user_id=sp.user_id).all()
                    if len(all_resps) == 1:
                        responses_by_step[(step_id, srv_id)] = all_resps[0]
                emitted_meta = set()
                for step_id, srv_id, q_id in q_cols:
                    step_key = (step_id, srv_id)
                    if step_key not in emitted_meta:
                        resp = responses_by_step.get(step_key)
                        row.append(resp.id if resp else '')
                        row.append(resp.completed_at.isoformat() if (resp and resp.completed_at) else '')
                        emitted_meta.add(step_key)
                    answers = responses_by_step.get(step_key)
                    row.append((answers.answers if answers else {}).get(str(q_id), ''))
                ws_sr.append(row)

            # Sheet: Task Submissions
            ws_sub = wb.create_sheet('Task_Submissions')
            sub_header = ['participant_id', 'condition', 'task_id', 'task_title',
                          'step_number', 'agent_id', 'agent_name',
                          'started_at', 'completed_at', 'time_spent_seconds',
                          'interactions', 'tokens_in', 'tokens_out',
                          'grade_value', 'grade_passed', 'grade_comment', 'submission_id']
            _ws_header(ws_sub, sub_header)
            task_step_map = {s.task_id: s for s in study.steps if s.step_type == 'task' and s.task_id}
            for sp in participants:
                subs = TaskSubmission.query.filter_by(study_id=study.id, user_id=sp.user_id).all()
                for sub in subs:
                    step = task_step_map.get(sub.task_id)
                    dur = ''
                    if sub.started_at and sub.completed_at:
                        dur = int((sub.completed_at - sub.started_at).total_seconds())
                    ws_sub.append([
                        _ident(sp), sp.condition.name if sp.condition else '',
                        sub.task_id, sub.task.title if sub.task else sub.task_id,
                        (step.step_order + 1) if step else '',
                        sub.agent_id or '', sub.agent_name or '',
                        sub.started_at.isoformat() if sub.started_at else '',
                        sub.completed_at.isoformat() if sub.completed_at else '',
                        dur, sub.interactions or 0,
                        sub.tokens_in or 0, sub.tokens_out or 0,
                        sub.grade_value if sub.grade_value is not None else '',
                        'pass' if sub.grade_passed else ('fail' if sub.grade_passed is False else ''),
                        _sanitize(sub.grade_comment or ''), sub.id,
                    ])

            # Sheet: Step Timings
            ws_t = wb.create_sheet('Step_Timings')
            _ws_header(ws_t, ['participant_id', 'step_number', 'step_label', 'step_type',
                               'started_at', 'completed_at', 'time_spent_seconds', 'is_late'])
            for sp in participants:
                for step in study.steps:
                    comp = StudyStepCompletion.query.filter_by(
                        participant_id=sp.id, step_id=step.id).first()
                    if comp:
                        ws_t.append([
                            _ident(sp),
                            step.step_order + 1, step.display_label, step.step_type,
                            comp.started_at.isoformat() if comp.started_at else '',
                            comp.completed_at.isoformat() if comp.completed_at else '',
                            comp.time_spent_seconds or '',
                            'yes' if comp.is_late else 'no',
                        ])

            # Sheet: Tracking Events (raw batches)
            ws_tr = wb.create_sheet('Tracking_Events')
            _ws_header(ws_tr, ['participant_id', 'task_id', 'submission_id',
                                'session_start', 'batch_seq', 'event_count', 'events_json'])
            for sp in participants:
                trackings = (TaskSessionTracking.query
                             .filter_by(study_id=study.id, user_id=sp.user_id)
                             .order_by(TaskSessionTracking.task_id, TaskSessionTracking.batch_seq)
                             .all())
                for tr in trackings:
                    try:
                        evts = json.loads(tr.events_data) if tr.events_data else []
                        ev_count = len(evts) if isinstance(evts, list) else 0
                    except Exception:
                        ev_count = 0
                    ws_tr.append([
                        _ident(sp), tr.task_id, tr.submission_id or '',
                        str(tr.session_start or ''), tr.batch_seq, ev_count,
                        tr.events_data or '[]',
                    ])

            # Sheet: BPMN Element Events (flattened)
            _BPMN_ETYPES = {'bpmn_add', 'bpmn_remove', 'bpmn_connect',
                            'bpmn_disconnect', 'bpmn_rename', 'bpmn_move'}
            ws_bpmn = wb.create_sheet('BPMN_Element_Events')
            _ws_header(ws_bpmn, [
                'participant_id', 'task_id', 'timestamp', 'event_type', 'source',
                'element_id', 'element_type', 'element_name',
                'source_element_id', 'source_element_name',
                'target_element_id', 'target_element_name',
                'x', 'y',
            ])
            for sp in participants:
                trackings = (TaskSessionTracking.query
                             .filter_by(study_id=study.id, user_id=sp.user_id)
                             .order_by(TaskSessionTracking.task_id, TaskSessionTracking.batch_seq)
                             .all())
                for tr in trackings:
                    try:
                        evts = json.loads(tr.events_data) if tr.events_data else []
                    except Exception:
                        evts = []
                    for ev in evts:
                        if not isinstance(ev, dict):
                            continue
                        if ev.get('type') not in _BPMN_ETYPES:
                            continue
                        ws_bpmn.append([
                            _ident(sp),
                            tr.task_id,
                            ev.get('ts', ''),
                            ev.get('type', ''),
                            ev.get('source', ''),
                            ev.get('element_id', ''),
                            ev.get('element_type', ''),
                            ev.get('element_name', ''),
                            ev.get('source_element_id', ''),
                            ev.get('source_element_name', ''),
                            ev.get('target_element_id', ''),
                            ev.get('target_element_name', ''),
                            ev.get('x', ''),
                            ev.get('y', ''),
                        ])

            # Sheet: LLM Interactions
            ws_llm = wb.create_sheet('LLM_Interactions')
            _ws_header(ws_llm, ['participant_id', 'task_id', 'submission_id',
                                 'call_index', 'timestamp', 'phase_label',
                                 'message_count', 'system_prompt', 'full_messages_json', 'response'])
            for sp in participants:
                subs = TaskSubmission.query.filter_by(study_id=study.id, user_id=sp.user_id).all()
                for sub in subs:
                    if not sub.llm_prompt_log:
                        continue
                    try:
                        calls = json.loads(sub.llm_prompt_log)
                    except Exception:
                        continue
                    if not isinstance(calls, list):
                        continue
                    for idx, call in enumerate(calls):
                        messages = call.get('messages', [])
                        sys_prompt = next((m.get('content', '') for m in messages if m.get('role') == 'system'), '')
                        ws_llm.append([
                            _ident(sp), sub.task_id, sub.id,
                            idx + 1, call.get('ts', ''), call.get('label', ''),
                            len(messages),
                            _sanitize(sys_prompt),
                            _sanitize(json.dumps(messages, ensure_ascii=False)),
                            _sanitize(call.get('response', '')),
                        ])

            # Sheet: Agent Choices
            ws_ac = wb.create_sheet('Agent_Choices')
            _ws_header(ws_ac, ['participant_id', 'wave_number', 'step_label',
                                'chosen_agent_id', 'chosen_agent_name', 'chosen_at'])
            for sp in participants:
                switches = (AgentSwitchHistory.query
                            .filter_by(participant_id=sp.id)
                            .order_by(AgentSwitchHistory.chosen_at)
                            .all())
                for sw in switches:
                    ws_ac.append([
                        _ident(sp), sw.wave_number,
                        sw.step.display_label if sw.step else '',
                        sw.agent_id or '',
                        sw.agent.name if sw.agent else '',
                        sw.chosen_at.isoformat() if sw.chosen_at else '',
                    ])

            # Save study workbook to ZIP
            wb_buf = io.BytesIO()
            wb.save(wb_buf)
            wb_buf.seek(0)
            zf.writestr(prefix + f'{slug}_data.xlsx', wb_buf.read())

            # BPMN files
            for sp in participants:
                subs = TaskSubmission.query.filter_by(study_id=study.id, user_id=sp.user_id).all()
                for sub in subs:
                    if sub.bpmn_xml:
                        fname = f'{prefix}bpmn/{_ident(sp)}_{sub.task_id}_sub{sub.id}.xml'
                        zf.writestr(fname, sub.bpmn_xml.encode('utf-8'))

    zip_buf.seek(0)
    safe_title = _slug(research.title)
    return Response(
        zip_buf.read(),
        mimetype='application/zip',
        headers={
            'Content-Disposition': (
                f'attachment; filename="research_{safe_title}_export_{ts}.zip"'
            )
        }
    )

