"""Admin — analytics and research data export routes."""
import io
import json
import zipfile
from datetime import datetime, timezone

from flask import render_template, request, send_file
from werkzeug.utils import secure_filename

from app.blueprints.admin import admin_bp
from app.models.user import User
from app.models.task import Task, TaskSubmission
from app.utils.decorators import admin_required, tutor_or_admin_required
from app.utils.audit import log_action
from app.utils.stats import (
    global_stats, since_from_period, chart_data_timeline,
    bpmn_error_frequency, phase_distribution, task_analytics, cohort_analytics,
)


@admin_bp.route('/analytics')
@tutor_or_admin_required
def analytics():
    period = request.args.get('period', '30d')
    since = since_from_period(period)
    gs = global_stats(since)
    chart = chart_data_timeline(None, since, period)
    errors = bpmn_error_frequency(since)
    phases = phase_distribution(since)
    tasks = task_analytics(since)
    cohorts = cohort_analytics(since)
    return render_template('cms/admin/analytics.html',
                           period=period, global_stats=gs, chart=chart,
                           bpmn_errors=errors, phases=phases,
                           task_analytics=tasks, cohort_analytics=cohorts)


@admin_bp.route('/export')
@admin_required
def export_page():
    return render_template('cms/admin/export.html')


@admin_bp.route('/export/generate', methods=['POST'])
@admin_required
def export_generate():
    import hashlib
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    export_format = request.form.get('export_format', 'zip')
    timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%d_%H%M%S')

    if export_format == 'json':
        def _anon(uid):
            if not uid:
                return None
            return hashlib.sha256(str(uid).encode()).hexdigest()[:16]

        subs = TaskSubmission.query.outerjoin(Task, TaskSubmission.task_id == Task.id).all()
        data = {
            'export_date': datetime.now(timezone.utc).isoformat(),
            'submissions': [
                {
                    'user_id_anon': _anon(s.user_id),
                    'task_id': s.task_id,
                    'task_title': s.task.title if s.task else s.task_id,
                    'started_at': s.started_at.isoformat() if s.started_at else None,
                    'completed_at': s.completed_at.isoformat() if s.completed_at else None,
                    'duration_seconds': round(s.duration_seconds, 1) if s.duration_seconds else None,
                    'interactions': s.interactions or 0,
                    'tokens_in': s.tokens_in or 0,
                    'tokens_out': s.tokens_out or 0,
                    'grade_value': s.grade_value,
                    'grade_passed': s.grade_passed,
                    'bpmn_xml': s.bpmn_xml,
                }
                for s in subs
            ],
        }
        buf = io.BytesIO(json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8'))
        buf.seek(0)
        log_action('export_research_data_json', 'System', None, {'timestamp': timestamp})
        return send_file(
            buf,
            mimetype='application/json',
            as_attachment=True,
            download_name=f'research_export_{timestamp}.json',
        )

    def _ws_header(ws, headers):
        ws.append(headers)
        hdr_fill = PatternFill('solid', fgColor='BDD7EE')
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold
            cell.fill = hdr_fill

    folder = f'export_{timestamp}'
    wb = Workbook()
    wb.remove(wb.active)

    # Submissions sheet
    ws_sub = wb.create_sheet('Submissions')
    _ws_header(ws_sub, ['id', 'user_id', 'username', 'task_id', 'task_title',
                         'started_at', 'completed_at', 'interactions', 'tokens_in',
                         'tokens_out', 'grade_value', 'grade_passed', 'grade_comment',
                         'ai_grade_value', 'ai_grade_passed'])
    for s in (TaskSubmission.query
              .outerjoin(User, TaskSubmission.user_id == User.id)
              .outerjoin(Task, TaskSubmission.task_id == Task.id)
              .all()):
        ws_sub.append([
            s.id, s.user_id,
            s.user.username if s.user else '',
            s.task_id,
            s.task.title if s.task else s.task_id,
            s.started_at.isoformat() if s.started_at else '',
            s.completed_at.isoformat() if s.completed_at else '',
            s.interactions or 0, s.tokens_in or 0, s.tokens_out or 0,
            s.grade_value, s.grade_passed, s.grade_comment or '',
            s.ai_grade_value, s.ai_grade_passed,
        ])

    # Users sheet
    ws_usr = wb.create_sheet('Users')
    _ws_header(ws_usr, ['id', 'username', 'email', 'role', 'created_at', 'is_active'])
    for u in User.query.order_by(User.id).all():
        ws_usr.append([u.id, u.username, u.email, u.role,
                       u.created_at.isoformat() if u.created_at else '',
                       u.is_active])

    # Tasks sheet
    ws_tsk = wb.create_sheet('Tasks')
    _ws_header(ws_tsk, ['id', 'title', 'grading_type', 'max_points', 'is_active', 'sort_order'])
    for t in Task.query.order_by(Task.sort_order).all():
        ws_tsk.append([t.id, t.title, t.grading_type, t.max_points,
                       t.is_active, t.sort_order])

    # Build ZIP with xlsx + bpmn + chat_logs
    xl_buf = io.BytesIO()
    wb.save(xl_buf)
    xl_buf.seek(0)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f'{folder}/data.xlsx', xl_buf.read())

        for s in TaskSubmission.query.filter(TaskSubmission.bpmn_xml.isnot(None)).all():
            username = s.user.username if s.user else 'unknown'
            fname = secure_filename(f'submission_{s.id}_{username}_{s.task_id}.bpmn')
            zf.writestr(f'{folder}/bpmn/{fname}', s.bpmn_xml or '')

        for s in TaskSubmission.query.filter(TaskSubmission.chat_log.isnot(None)).all():
            fname = f'submission_{s.id}.json'
            zf.writestr(f'{folder}/chat_logs/{fname}', s.chat_log or '[]')

    zip_buf.seek(0)
    log_action('export_research_data', 'System', None, {'timestamp': timestamp})
    return send_file(
        zip_buf,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'{folder}.zip',
    )
